import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import tkinter as tk
from tkinter import filedialog
from helper import resource_path

# Define your sheet name
SheetName = "Input Form"

# Define the mapping of JSON keys to cell locations
cell_mapping = {
    'no': (SheetName, 'J3'),
    'kep_keluarga': (SheetName, 'D4'),
    'kelurahan': (SheetName, 'Q4'),
    'alamat': (SheetName, 'D5'),
    'kecamatan': (SheetName, 'Q5'),
    'rw': (SheetName, 'D6'),
    'kota': (SheetName, 'Q6'),
    'pos': (SheetName, 'D7'),
    'provinsi': (SheetName, 'Q7'),
    'keluarga': (SheetName, 'H39'),
    'officer_name': (SheetName, 'N39'),
    'nip': (SheetName, 'O40'),

    # Split a single string like "23-04-2021" into 3 cells: D34, E34, F34
    'tanggal': {'sheet': SheetName, 'cell': 'D34', 'split': '-'},

    # Array fields (e.g., list of names)
    'names': (SheetName, 'B10'),
    'niks': (SheetName, 'F10'),
    'sexes': (SheetName, 'H10'),
    'birthplaces': (SheetName, 'J10'),
    'religions': (SheetName, 'N10'),
    'educations': (SheetName, 'P10'),
    'profession': (SheetName, 'S10'),
    'marriage_stats': (SheetName, 'B24'),
    'marriage_dates': ("Sheet1", 'C24'),

    # Array of date strings, each like "23-04-2021", needs to be split into K-row, L-row, M-row
    'birthdates': {'sheet': SheetName, 'cell': 'K10', 'split': '-', 'is_array_of_strings': True},

    'marriage_rels': (SheetName, 'D24'),
    'citizenships': (SheetName, 'F24'),
    'paspor_no': (SheetName, 'H24'),
    'kitas_no': (SheetName, 'K24'),
    'father_names': (SheetName, 'N24'),
    'mother_names': (SheetName, 'Q24'),
}

def populate_excel(workbook_path, input_json_path, final_output_path):
    # Load your Excel file
    workbook = load_workbook(resource_path(workbook_path))

    # Load your JSON file
    with open(input_json_path, encoding='utf-8') as f:
        data = json.load(f)

    # Fill in the Excel based on the mapping
    for key, config in cell_mapping.items():
        if key in data:
            value = data[key]

            # If using dict-style mapping (special handling like split)
            if isinstance(config, dict):
                sheet_name = config['sheet']
                start_cell = config['cell']
                split = config.get('split')
                is_array_of_strings = config.get('is_array_of_strings', False)

                ws = workbook[sheet_name]  # Get the sheet

                if is_array_of_strings:
                    # Array of strings to split across columns
                    col = column_index_from_string(start_cell[:1])
                    row = int(start_cell[1:])
                    for item in value:
                        split_values = item.split(split)
                        for offset, val in enumerate(split_values):
                            try:
                                num_val = int(val)
                            except ValueError:
                                num_val = val
                            ws.cell(row=row, column=col + offset, value=num_val)
                        row += 1

                elif isinstance(value, str) and split:
                    # Single string split across columns
                    split_values = value.split(split)
                    col = column_index_from_string(start_cell[:1])
                    row = int(start_cell[1:])
                    for offset, val in enumerate(split_values):
                        try:
                            num_val = int(val)
                        except ValueError:
                            num_val = val
                        ws.cell(row=row, column=col + offset, value=num_val)

            # If using tuple-style mapping (normal mapping or array to column)
            elif isinstance(config, tuple):
                sheet_name, start_cell = config
                ws = workbook[sheet_name]

                if isinstance(value, list):
                    # Array: fill down a column starting at the specified cell
                    col_letter = start_cell[:1]
                    row = int(start_cell[1:])
                    for item in value:
                        ws[f'{col_letter}{row}'] = item
                        row += 1
                else:
                    # Single value
                    ws[start_cell] = value
                    
    if not final_output_path:
        # Hide the root window
        root = tk.Tk()
        root.withdraw()

        # Open the Save As dialog
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save Excel File As"
        )

        # Save if a file was selected
        if file_path:
            workbook.save(file_path)
            print(f"Workbook saved to: {file_path}")
        else:
            print("Save cancelled.")
    else :
        workbook.save(final_output_path)